from ..models import Job, Task, Patient, Configuration
from .SendEmail import SendEmail
from django.contrib.auth.models import User
from datetime import date
from django.conf import settings
from django.db import connection
from django.contrib import messages
from django.core.files import File
import openpyxl  #package to open an excel file
import os

class DatabaseOperations:
    """Functions that directly operate on the database"""

    def populate_Configuration_table(self):
        """
        The first time a user navigates to the home page, this function
        is executed in order to populate the Configuration table with 
        initial values. 
        """
        try:
            if Configuration.objects.count() == 0:
               #empty table, so populate it with initial values
               first_table_row = Configuration(main_title = 'initial value',
                    main_intro = 'initial value',
                    indiv_intro = 'initial value',
                    number_days_to_complete=7,
                    max_num_jobs=4)
               first_table_row.save()
        except Exception as e:
            return HttpResponse("Error in DatabaseOperations.populate_Configuration_table {}".format(e))


    def deadline_passed_set_job_available(self):
        """
        This function sets a job's status back to 'Available' when 
        its status is still 'In Progress' and its deadline date has passed.
        """
        try:
            jobs = Job.objects.filter(deadline_date__lt=date.today(), status ='In Progress')
            jobs.update(status ='Available',
                        user_id = None,
                        start_date = None,
                        deadline_date = None,
                        reminder_sent= 'no')
        except Exception as e:
            return HttpResponse("Error in DatabaseOperations.deadline_passed_set_job_available {}".format(e))


    def clear_database(self):
        """Delete all data in the database to create empty tables"""
        try:
            Job.objects.all().delete()
            Patient.objects.all().delete()
            Task.objects.all().delete()
            #Reset the auto-increment primary key id field back to an initial value of 1
            cursor = connection.cursor()
            cursor.execute("ALTER Sequence jobs_job_id_seq RESTART with 1;")
            cursor.execute("ALTER Sequence jobs_task_task_id_seq RESTART with 1;")
            connection.commit()
            connection.close()
        except Exception as e:
            return HttpResponse("Error in DatabaseOperations.clear_database {}".format(e))


    def populate_task_table(self, wb):
        """Populate the Task table using the data in an Excel spreadsheet"""
        try:
            wsTask = wb["Tasks"]
            # iterating over the rows and
            # getting value from each cell in row
            for row_num, row in enumerate(wsTask.iter_rows()):
                #Ignore header row
                if row_num > 0:
                    new_task_name =row[0].value
                    if row[1].value is None:
                        new_repetitions= 1
                    else:
                        new_repetitions=row[1].value

                    task, _ = Task.objects.update_or_create(task_name=new_task_name, repetitions=new_repetitions)
                    task.save()
        except Exception as e:
            return HttpResponse("Error in DatabaseOperations.populate_task_table {}".format(e))


    def populate_patient_table(self, wb):
        """Populate the Patient table using the data in an Excel spreadsheet"""
        try:
            wsPatient = wb["Patients"]
            # iterating over the rows and
            # getting value from each cell in row
            for row_num, row in enumerate(wsPatient.iter_rows()):
                #Ignore header row
                if row_num > 0:
                    patient, _ = Patient.objects.update_or_create(patient_id =row[0].value)
                    patient.save()
        except Exception as e:
            return HttpResponse("Error in DatabaseOperations.populate_patient_table {}".format(e))


    def populate_job_table(self):
        """Populate the Job table using the data in the Task & Patient tables"""
        try:
            patient_list = Patient.objects.all()
            task_list = Task.objects.all()
            task_counter = 1
            for patient in patient_list:
                for task in task_list:
                    num_repetions = task.repetitions
                    for i in range(num_repetions):
                        job = Job(patient_id=patient, task_id=task, repetition_num=i+1)
                        job.save()
        except Exception as e:
            return HttpResponse("Error in DatabaseOperations.populate_job_table {}".format(e))


    def populate_database_from_file(self):
        """ This function populates the database from a local file.
            It is used to populate a database for testing
        """
        file_path = settings.BASE_DIR + '/Test_Data_Do_Not_Delete.xlsx'
        with open(file_path, 'rb') as file:
            excel_file = File(file)
            wb = openpyxl.load_workbook(excel_file)
            #self.clear_database()
            self.populate_task_table(wb)
            self.populate_patient_table(wb)
            #create job table
            self.populate_job_table()



    def populate_database(self, request):
        """Deletes the contents of the database and then repopulates it"""
        try:
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            self.clear_database()
            self.populate_task_table(wb)
            self.populate_patient_table(wb)
            #create job table
            self.populate_job_table()
        except Exception as e:
           messages.error(request,"Error in views.dbAdmin.populate_database opening uploaded Excel file.")
                
        